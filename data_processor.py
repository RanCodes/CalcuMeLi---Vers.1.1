import pandas as pd
from utils import (
    parse_fee_combo,
    parse_pct,
    parse_money,
    clean_ml_data,
    validate_excel_structure,
    extract_tax_percentage,
    calcular_precio_publicacion_ml,
)

def leer_ml(file_path_or_buffer) -> pd.DataFrame:
    """
    Lee el archivo Excel de MercadoLibre y lo limpia.

    Args:
        file_path_or_buffer: Ruta al archivo o buffer de bytes

    Returns:
        DataFrame limpio con datos válidos de ML
    """
    try:
        # Intentar leer la hoja "Hoja1"
        try:
            df = pd.read_excel(file_path_or_buffer, sheet_name='Hoja1')
        except ValueError:
            # Si no existe "Hoja1", usar la primera hoja
            df = pd.read_excel(file_path_or_buffer, sheet_name=0)

        # Validar estructura
        is_valid, error_msg = validate_excel_structure(df, 'ml')
        if not is_valid:
            raise ValueError(f"Error en estructura ML: {error_msg}")

        # Limpiar datos
        df_clean = clean_ml_data(df)

        # Parsear campos específicos
        df_clean['fee_pct'], df_clean['fee_fixed'] = zip(*df_clean['FEE_PER_SALE_MARKETPLACE_V2'].apply(parse_fee_combo))
        df_clean['financing_pct'] = df_clean['COST_OF_FINANCING_MARKETPLACE'].apply(parse_pct)

        # Convertir tipos de datos
        df_clean['PRICE'] = df_clean['PRICE'].apply(parse_money)
        df_clean['QUANTITY'] = pd.to_numeric(df_clean['QUANTITY'], errors='coerce').fillna(0)

        return df_clean

    except Exception as e:
        raise Exception(f"Error al leer archivo MercadoLibre: {str(e)}")

def leer_odoo(file_path_or_buffer) -> pd.DataFrame:
    """
    Lee el archivo Excel de Odoo.

    Args:
        file_path_or_buffer: Ruta al archivo o buffer de bytes

    Returns:
        DataFrame con datos de Odoo
    """
    try:
        # Intentar leer la hoja "Sheet1"
        try:
            df = pd.read_excel(file_path_or_buffer, sheet_name='Sheet1')
        except ValueError:
            # Si no existe "Sheet1", usar la primera hoja
            df = pd.read_excel(file_path_or_buffer, sheet_name=0)

        # Validar estructura
        is_valid, error_msg = validate_excel_structure(df, 'odoo')
        if not is_valid:
            raise ValueError(f"Error en estructura Odoo: {error_msg}")

        # Limpiar datos
        # Filtrar filas donde al menos el código no esté vacío
        df_clean = df[df['Código Neored'].notna()].copy()

        # Convertir tipos de datos
        df_clean['Precio Tarifa'] = pd.to_numeric(df_clean['Precio Tarifa'], errors='coerce').fillna(0)
        df_clean['Cantidad a mano'] = pd.to_numeric(df_clean['Cantidad a mano'], errors='coerce').fillna(0)

        # Parsear porcentaje de impuestos
        df_clean['tax_pct'] = df_clean['Impuestos del cliente'].apply(extract_tax_percentage)

        return df_clean

    except Exception as e:
        raise Exception(f"Error al leer archivo Odoo: {str(e)}")

def unir_y_validar(df_ml: pd.DataFrame, df_odoo: pd.DataFrame) -> pd.DataFrame:
    """
    Une los DataFrames de ML y Odoo por SKU y valida el resultado.

    Args:
        df_ml: DataFrame de MercadoLibre
        df_odoo: DataFrame de Odoo

    Returns:
        DataFrame unido con flags de validación
    """
    # Hacer join por SKU
    df_merged = df_ml.merge(
        df_odoo,
        left_on='SKU',
        right_on='Código Neored',
        how='left',
        suffixes=('_ml', '_odoo')
    )

    # Crear columna de flags/notas
    df_merged['Notas/Flags'] = ''

    # Validar matcheo
    no_match_mask = df_merged['Código Neored'].isna()
    df_merged.loc[no_match_mask, 'Notas/Flags'] = 'SKU no encontrado en Odoo'

    # Validar datos críticos
    missing_price_mask = (df_merged['Precio Tarifa'].isna()) | (df_merged['Precio Tarifa'] == 0)
    df_merged.loc[missing_price_mask, 'Notas/Flags'] += '; Precio Tarifa faltante'

    # Validar stock
    missing_stock_mask = df_merged['Cantidad a mano'].isna()
    df_merged.loc[missing_stock_mask, 'Notas/Flags'] += '; Stock faltante'

    # Limpiar flags vacías
    df_merged['Notas/Flags'] = df_merged['Notas/Flags'].str.strip('; ')

    return df_merged

def calcular(
    df: pd.DataFrame,
    base_financiacion: str = 'tarifa',
    incluir_impuestos: bool = False,
    tipo_recargo_envio: str = 'Ninguno',
    valor_recargo_envio: float = 0.0
) -> pd.DataFrame:
    """
    Calcula los precios finales con el desglose de recargos.

    Args:
        df: DataFrame unido y validado
        base_financiacion: 'tarifa' o 'tarifa_mas_ml'
        incluir_impuestos: Si incluir impuestos del cliente en la tarifa
        tipo_recargo_envio: 'Ninguno', 'Fijo ($)' o 'Porcentaje (%)'
        valor_recargo_envio: Monto fijo o porcentaje según corresponda

    Returns:
        DataFrame con cálculos completados
    """
    df_calc = df.copy()

    if 'Notas/Flags' not in df_calc.columns:
        df_calc['Notas/Flags'] = ''
    df_calc['Notas/Flags'] = df_calc['Notas/Flags'].fillna('')

    df_calc['Precio de Tarifa'] = pd.to_numeric(
        df_calc['Precio Tarifa'], errors='coerce'
    ).fillna(0.0)

    tax_pct = pd.to_numeric(df_calc.get('tax_pct', 0.0), errors='coerce').fillna(0.0)
    tarifa_base = df_calc['Precio de Tarifa']
    tarifa_con_impuestos = tarifa_base * (1 + tax_pct)
    if incluir_impuestos:
        tarifa_neta_base = tarifa_con_impuestos
        df_calc['Tarifa + impuestos'] = tarifa_con_impuestos
    else:
        tarifa_neta_base = tarifa_base
        df_calc['Tarifa + impuestos'] = tarifa_con_impuestos

    # Inicializar columnas
    df_calc['Recargo % ML (importe)'] = 0.0
    df_calc['Recargo fijo ML ($)'] = 0.0
    df_calc['Cargo por vender ($)'] = 0.0
    df_calc['Recargo financiación (importe)'] = 0.0
    df_calc['Recargo envío ($)'] = 0.0
    df_calc['Retenciones ML ($)'] = 0.0
    df_calc['Recibis ($)'] = 0.0
    df_calc['IVA'] = 0.0
    df_calc['Precio final'] = 0.0

    # Identificar filas a las que se les debe aplicar recargo de envío
    shipping_column = next(
        (col for col in ['SHIPPING_METHOD ', 'SHIPPING_METHOD'] if col in df_calc.columns),
        None
    )
    if shipping_column:
        shipping_values = df_calc[shipping_column].fillna('').astype(str)
        aplica_envio = shipping_values.str.contains(
            'Mercado Envíos por mi cuenta',
            case=False,
            regex=False
        )
    else:
        aplica_envio = pd.Series(False, index=df_calc.index, dtype=bool)

    # Calcular recargo de envío solo para las filas aplicables
    tipo_envio = tipo_recargo_envio.lower() if isinstance(tipo_recargo_envio, str) else 'ninguno'
    if tipo_envio.startswith('fijo') and valor_recargo_envio:
        try:
            monto_fijo = float(valor_recargo_envio)
        except (TypeError, ValueError):
            monto_fijo = 0.0
        if monto_fijo != 0.0:
            df_calc.loc[aplica_envio, 'Recargo envío ($)'] = monto_fijo
        else:
            df_calc.loc[aplica_envio, 'Recargo envío ($)'] = 0.0
    elif tipo_envio.startswith('porcentaje') and valor_recargo_envio:
        try:
            pct_envio = float(valor_recargo_envio)
        except (TypeError, ValueError):
            pct_envio = 0.0
        if pct_envio > 1:
            pct_envio = pct_envio / 100.0
        df_calc.loc[aplica_envio, 'Recargo envío ($)'] = tarifa_neta_base[aplica_envio] * pct_envio

    tarifa_objetivo = tarifa_neta_base + df_calc['Recargo envío ($)']

    fee_pct = pd.to_numeric(df_calc.get('fee_pct', 0.0), errors='coerce').fillna(0.0)
    fee_fixed = pd.to_numeric(df_calc.get('fee_fixed', 0.0), errors='coerce').fillna(0.0)
    financing_pct = pd.to_numeric(df_calc.get('financing_pct', 0.0), errors='coerce').fillna(0.0)
    if 'retenciones_pct' in df_calc.columns:
        retenciones_pct = pd.to_numeric(df_calc['retenciones_pct'], errors='coerce').fillna(0.0)
    else:
        retenciones_pct = pd.Series(0.0, index=df_calc.index)

    calculos = [
        calcular_precio_publicacion_ml(
            tarifa_neta=tarifa_objetivo.iat[idx],
            porcentaje_comision=fee_pct.iat[idx],
            porcentaje_financiacion=financing_pct.iat[idx],
            porcentaje_retenciones=retenciones_pct.iat[idx],
            costo_fijo=fee_fixed.iat[idx],
        )
        for idx in range(len(df_calc))
    ]

    calculos_df = pd.DataFrame(
        calculos,
        columns=[
            'Precio final',
            'Cargo por vender ($)',
            'Recargo financiación (importe)',
            'Retenciones ML ($)',
            'Recibis ($)',
            'Denominador inválido',
        ],
        index=df_calc.index,
    )

    df_calc['Precio final'] = calculos_df['Precio final']
    df_calc['Cargo por vender ($)'] = calculos_df['Cargo por vender ($)']
    df_calc['Recargo financiación (importe)'] = calculos_df['Recargo financiación (importe)']
    df_calc['Retenciones ML ($)'] = calculos_df['Retenciones ML ($)']
    df_calc['Recibis ($)'] = calculos_df['Recibis ($)']
    df_calc['Recargo fijo ML ($)'] = fee_fixed
    df_calc['Recargo % ML (importe)'] = df_calc['Precio final'] * fee_pct

    invalid_mask = calculos_df['Denominador inválido']
    if invalid_mask.any():
        mensaje = 'Porcentajes ML sin solución (denominador <= 0)'
        df_calc.loc[invalid_mask, 'Notas/Flags'] = df_calc.loc[invalid_mask, 'Notas/Flags'].apply(
            lambda x: mensaje if not x else f"{x}; {mensaje}"
        )
        df_calc.loc[invalid_mask, 'Recargo fijo ML ($)'] = 0.0

    df_calc['IVA'] = 0.0
    positive_tax = tax_pct > 0
    if positive_tax.any():
        df_calc.loc[positive_tax, 'IVA'] = (
            df_calc.loc[positive_tax, 'Precio final']
            * tax_pct[positive_tax]
            / (1 + tax_pct[positive_tax])
        )

    # Redondear a 2 decimales
    numeric_cols = [
        'Precio de Tarifa',
        'Tarifa + impuestos',
        'Recargo % ML (importe)',
        'Recargo fijo ML ($)',
        'Cargo por vender ($)',
        'Recargo financiación (importe)',
        'Recargo envío ($)',
        'Retenciones ML ($)',
        'Recibis ($)',
        'IVA',
        'Precio final',
    ]
    for col in numeric_cols:
        df_calc[col] = df_calc[col].round(2)

    # Preparar columnas adicionales
    df_calc['% ML aplicado'] = (fee_pct * 100).round(2)
    df_calc['% financiación aplicado'] = (financing_pct * 100).round(2)

    return df_calc

def preparar_resultado_final(
    df_calc: pd.DataFrame,
    incluir_impuestos: bool = False,
    incluir_envio: bool = False
) -> pd.DataFrame:
    """
    Prepara el DataFrame final con las columnas en el orden exacto especificado.

    Args:
        df_calc: DataFrame con cálculos completados
        incluir_impuestos: Si se incluyeron impuestos en el cálculo
        incluir_envio: Si se incluyó recargo de envío en el cálculo

    Returns:
        DataFrame con estructura final para exportar
    """
    # Definir columnas fijas
    columnas_finales = [
        'Numero de publicación',
        'SKU',
        'Descripción del producto',
        'Stock',
        'Precio de Tarifa'
    ]

    # Agregar tarifa + impuestos si procede
    if incluir_impuestos:
        columnas_finales.append('Tarifa + impuestos')

    # Columnas comunes de resultados
    columnas_finales.extend([
        'Precio final',
        'IVA',
        'Recargo % ML (importe)',
        'Recargo fijo ML ($)',
        'Cargo por vender ($)',
        'Recargo financiación (importe)',
        'Retenciones ML ($)',
        'Recibis ($)'
    ])

    # Añadir recargo de envío si corresponde
    if incluir_envio:
        columnas_finales.append('Recargo envío ($)')

    columnas_finales.extend([
        '% ML aplicado',
        '% financiación aplicado',
        'Tipo de publicación',
        'Precio actual en ML',
        'Moneda'
    ])

    # Agregar columna de notas al final
    columnas_finales.append('Notas/Flags')

    # Crear DataFrame resultado
    df_resultado = pd.DataFrame()

    # Mapear campos
    df_resultado['Numero de publicación'] = df_calc['ITEM_ID']
    df_resultado['SKU'] = df_calc['SKU']
    df_resultado['Descripción del producto'] = df_calc['Nombre'].fillna(df_calc['TITLE'])
    df_resultado['Stock'] = df_calc['Cantidad a mano'].fillna(0).astype(int)
    df_resultado['Precio de Tarifa'] = df_calc['Precio de Tarifa']

    if incluir_impuestos:
        df_resultado['Tarifa + impuestos'] = df_calc['Tarifa + impuestos']

    df_resultado['Precio final'] = df_calc['Precio final']
    df_resultado['IVA'] = df_calc['IVA']
    df_resultado['Recargo % ML (importe)'] = df_calc['Recargo % ML (importe)']
    df_resultado['Recargo fijo ML ($)'] = df_calc['Recargo fijo ML ($)']
    df_resultado['Cargo por vender ($)'] = df_calc['Cargo por vender ($)']
    df_resultado['Recargo financiación (importe)'] = df_calc['Recargo financiación (importe)']
    df_resultado['Retenciones ML ($)'] = df_calc['Retenciones ML ($)']
    df_resultado['Recibis ($)'] = df_calc['Recibis ($)']
    if incluir_envio:
        df_resultado['Recargo envío ($)'] = df_calc['Recargo envío ($)']
    df_resultado['% ML aplicado'] = df_calc['% ML aplicado']
    df_resultado['% financiación aplicado'] = df_calc['% financiación aplicado']
    df_resultado['Tipo de publicación'] = df_calc['LISTING_TYPE_V3']
    df_resultado['Precio actual en ML'] = df_calc['PRICE']
    df_resultado['Moneda'] = df_calc['CURRENCY_ID']
    df_resultado['Notas/Flags'] = df_calc['Notas/Flags'].fillna('')

    # Reordenar columnas
    df_resultado = df_resultado[columnas_finales]

    return df_resultado

def exportar_excel(df: pd.DataFrame, output_path: str = None) -> bytes:
    """
    Exporta el DataFrame a Excel y retorna los bytes del archivo.

    Args:
        df: DataFrame a exportar
        output_path: Ruta opcional para guardar archivo

    Returns:
        bytes: Contenido del archivo Excel
    """
    import io
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "resultado"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
        for cell in column[1:]:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if output_path:
        with open(output_path, 'wb') as f:
            f.write(buffer.getvalue())
        buffer.seek(0)

    return buffer.getvalue()